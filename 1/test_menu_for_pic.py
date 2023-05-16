import pandas as pd
from openpyxl import Workbook
import shutil
import openpyxl
import time

# 엑셀 파일 읽기

df = pd.read_excel('D:\\for_project_files\\yami.xlsx',engine='openpyxl')


# 조식, 중식, 석식 구분값 추출
values1 = df.iloc[1, 0]
values2 = df.iloc[14, 0]
values3 = df.iloc[18, 0]

# 해당 값들을 DataFrame으로 변환 (열 이름은 기존 데이터와 동일하게 설정)
values_df1 = pd.DataFrame({df.columns[0]: [values1]})
values_df2 = pd.DataFrame({df.columns[0]: [values2]})
values_df3 = pd.DataFrame({df.columns[0]: [values3]})

# 각 요일별로 저장
for i in range(1, 6):
    shutil.copy('D:\\for_project_files\\template.xlsx', f'D:\\for_project_files\\new_yami{i}.xlsx')
    wb = openpyxl.load_workbook(f'D:\\for_project_files\\new_yami{i}.xlsx')
    ws = wb.active
    # 범위 추출.
    day = df.iloc[0, i]
    df1 = df.iloc[1:7, i:i+1]   # 첫 번째 열('A'열)의 3행부터 7행까지 값
    df2 = df.iloc[14:18, i:i+1] # 첫 번째 열('A'열)의 18행부터 20행까지 값
    df3 = df.iloc[18:25, i:i+1] # 첫 번째 열('A'열)의 22행부터 26행까지 값

    # 새 워크북 생성
    wb = Workbook()

    ws['A3'] = values1
    ws['B1'] = day  # 요일 값
    for idx, val in enumerate(df1.iloc[:, 0].values):
        ws[f'B{idx+4}'] = val[0:]  # df1의 나머지 값들

    ws['A10'] = values2
    for idx, val in enumerate(df2.iloc[:, 0].values):
        ws[f'B{idx+11}'] = val  # df2의 값들

    ws['A15'] = values3
    for idx, val in enumerate(df3.iloc[:, 0].values):
        ws[f'B{idx+16}'] = val  # df3의 값들

    wb.save(f'D:/for_project_files/new_yami{i}.xlsx')
time.sleep(1)
# for i in range(1, 6):
#     shutil.copy('../../../for_project_files/template.xlsx', f'new_yami{i}.xlsx')

# # 복사한 파일에 데이터 추가
#     wb = openpyxl.load_workbook(f'D:/for_project_files/new_yami{i}.xlsx')
#     ws = wb.active
#
#     # 범위 추출.
#     day = df.iloc[0, i]
#     unmerged_cell = ws['B1']
#     unmerged_cell.value = day
#     df1 = df.iloc[1:7, i:i+1]   # 첫 번째 열('A'열)의 3행부터 7행까지 값
#     df2 = df.iloc[14:18, i:i+1] # 첫 번째 열('A'열)의 18행부터 20행까지 값
#     df3 = df.iloc[18:25, i:i+1] # 첫 번째 열('A'열)의 22행부터 26행까지 값
#
#     ws['A3'] = values1
#
#     ws['B1'] = day  # 요일 값
#     for idx, val in enumerate(df1.iloc[:, 0].values):
#         ws[f'B{idx+4}'] = val[0:]  # df1의 나머지 값들
#
#     ws['A10'] = values2
#     for idx, val in enumerate(df2.iloc[:, 0].values):
#         ws[f'B{idx+11}'] = val  # df2의 값들
#
#     ws['A15'] = values3
#     for idx, val in enumerate(df3.iloc[:, 0].values):
#         ws[f'B{idx+16}'] = val  # df3의 값들
#
#     wb.save(f'D:/for_project_files/fin_new_yami{i}.xlsx')

# DataFrame을 하나로 결합
    # new_df = pd.concat([values_df1, df1, values_df2, df2, values_df3, df3])
    #
    # # 새로운 엑셀 파일로 저장
    # new_df.to_excel(f'D:/for_project_files/new_yami{i}.xlsx', index=False)
